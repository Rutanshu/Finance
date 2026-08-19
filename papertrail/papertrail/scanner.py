"""
PaperTrail :: scanner
Walks a local folder of Excel workbooks and extracts the dependency structure.

Everything here is offline. No network calls, no telemetry, no uploads.
The only I/O is: read .xlsx/.xlsm files, write a JSON file next to them.
"""

from __future__ import annotations

import os
import re
import zipfile
from dataclasses import dataclass, field, asdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

EXCEL_EXT = {".xlsx", ".xlsm", ".xltx", ".xltm"}
SKIP_PREFIX = ("~$", ".")
MAX_HEADER_COLS = 60
MAX_LINKS_PER_BOOK = 4000

# ='[Revenue_Register.xlsx]Data'!$B$4   or   =[1]Sheet1!A1   or   ='C:\path\[Book.xlsx]Sheet'!A1
RE_EXTERNAL = re.compile(r"\[([^\[\]]+?\.xls[xmb]?)\]", re.IGNORECASE)
RE_INDEXED = re.compile(r"\[(\d+)\]")
# 'Sheet Name'!A1  or  SheetName!A1
RE_SHEETREF = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_. ]*))!\$?[A-Z]{1,3}\$?\d+")
RE_VOLATILE = re.compile(r"\b(INDIRECT|OFFSET|NOW|TODAY|RAND|RANDBETWEEN)\s*\(", re.IGNORECASE)
RE_LOOKUP = re.compile(r"\b(VLOOKUP|HLOOKUP|XLOOKUP|INDEX|MATCH)\s*\(", re.IGNORECASE)
# a formula that reaches out live at calculation time: WEBSERVICE(url), FILTERXML(...), RTD(...) real-time-data feeds
RE_WEBFORMULA = re.compile(r"\b(WEBSERVICE|FILTERXML|RTD)\s*\(", re.IGNORECASE)
RE_URL = re.compile(r'"(https?://[^"]+)"', re.IGNORECASE)

# Power Query / Get&Transform "M" source functions — whatever they're pulling from, it isn't this folder.
# Web.*/OData.Feed -> a website or web API. Sql./Odbc./OleDb. -> a database. SharePoint./Exchange. -> a cloud service.
RE_M_SOURCE = re.compile(
    r"\b(Web\.Contents|Web\.Page|OData\.Feed|Sql\.Database|Sql\.Databases|Odbc\.DataSource|Odbc\.Query|"
    r"OleDb\.DataSource|SharePoint\.Contents|SharePoint\.Files|Exchange\.Contents|Facebook\.Contents|"
    r"Folder\.Files|Folder\.Contents)\s*\(\s*\"?([^\",)]*)",
    re.IGNORECASE,
)
M_SOURCE_KIND = {
    "web.contents": "web_query", "web.page": "web_query", "odata.feed": "web_query",
    "sql.database": "database", "sql.databases": "database", "odbc.datasource": "database",
    "odbc.query": "database", "oledb.datasource": "database",
    "sharepoint.contents": "sharepoint", "sharepoint.files": "sharepoint",
    "exchange.contents": "exchange", "facebook.contents": "web_query",
    "folder.files": "local_folder", "folder.contents": "local_folder",
}

# cell-level external reference, quoted form:  '[Revenue_Register.xlsx]Data'!$B$4  or  a range like !$B$2:$B$10
# and unquoted form (sheet name has no spaces):  [Revenue_Register.xlsx]Data!$B$4
RE_EXTCELL = re.compile(
    r"'\[(?P<f1>[^\[\]]+?\.xls[xmb]?)\](?P<s1>[^'!\[\]]*)'!(?P<c1>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)"
    r"|\[(?P<f2>[^\[\]]+?\.xls[xmb]?)\](?P<s2>[A-Za-z_][A-Za-z0-9_.]*)!(?P<c2>\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)",
    re.IGNORECASE,
)


@dataclass
class SheetInfo:
    name: str
    rows: int = 0
    cols: int = 0
    formula_cells: int = 0
    constant_cells: int = 0
    hardcoded_in_formula: int = 0       # =1234+A1  -> magic numbers buried in formulas
    volatile_cells: int = 0
    lookup_cells: int = 0
    web_formula_cells: int = 0          # WEBSERVICE/FILTERXML/RTD -> reaches the internet at calc time
    headers: list[str] = field(default_factory=list)


@dataclass
class BookInfo:
    id: str
    name: str
    path: str
    rel_path: str
    size_kb: int
    modified: str
    sheets: list[SheetInfo] = field(default_factory=list)
    external_refs: list[str] = field(default_factory=list)   # raw workbook names it points at
    unresolved_refs: list[str] = field(default_factory=list)  # pointed at, but not found in folder
    links: list[dict] = field(default_factory=list)          # cell-level cross-file references, see _cell_link()
    web_links: list[dict] = field(default_factory=list)      # cell-level WEBSERVICE/FILTERXML/RTD formulas
    external_sources: list[dict] = field(default_factory=list)  # connections.xml / Power Query sources: db, web, etc.
    error: str | None = None

    @property
    def formula_count(self) -> int:
        return sum(s.formula_cells for s in self.sheets)


def _book_id(rel_path: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", rel_path.lower()).strip("_")


def _leading_col(cell_ref: str) -> str:
    """'$B$4' or 'B4:B10' -> 'B'"""
    m = re.match(r"\$?([A-Z]{1,3})", cell_ref)
    return m.group(1) if m else ""


def _external_link_targets(path: Path) -> list[str]:
    """
    openpyxl does not always surface externalLink targets, so read them
    straight out of the .xlsx zip container. Purely a local file read.
    """
    out: list[str] = []
    try:
        with zipfile.ZipFile(path) as z:
            for item in z.namelist():
                if item.startswith("xl/externalLinks/_rels/"):
                    raw = z.read(item).decode("utf-8", "ignore")
                    for m in re.finditer(r'Target="([^"]+)"', raw):
                        target = m.group(1)
                        if target.lower().startswith(("http://", "https://")):
                            out.append(target)          # keep full URL, don't basename it away
                        else:
                            out.append(os.path.basename(target.replace("\\", "/")))
    except (zipfile.BadZipFile, KeyError, OSError):
        pass
    return out


def _redact(conn_str: str) -> str:
    """Strip credential-shaped tokens (Pwd=..., Password=..., User ID=...) out of an ODBC/OLEDB connection string."""
    conn_str = re.sub(r"(?i)\b(pwd|password|uid|user\s*id)\s*=\s*[^;]*", r"\1=•••", conn_str)
    return conn_str[:160]


def _friendly_conn_detail(conn_str: str) -> str:
    conn_str = _redact(conn_str)
    for key in ("Data Source", "Server", "Host", "DSN", "Driver", "Database"):
        m = re.search(rf"(?i)\b{re.escape(key)}\s*=\s*([^;]+)", conn_str)
        if m:
            return f"{key}={m.group(1).strip()}"
    return conn_str or "(unlabelled connection)"


def _attr(tag: str, name: str) -> str | None:
    """Pull one attribute out of an XML start-tag, regardless of what order the attributes appear in."""
    m = re.search(rf'\b{re.escape(name)}="([^"]*)"', tag)
    return m.group(1) if m else None


def _connection_sheet_map(z: zipfile.ZipFile) -> dict[str, str]:
    """
    Best-effort: which worksheet does each connection id load into?
    Chain:  xl/workbook.xml (sheet name -> r:id) + xl/_rels/workbook.xml.rels (r:id -> worksheets/sheetN.xml)
            -> xl/worksheets/_rels/sheetN.xml.rels (-> queryTables/queryTableM.xml)
            -> queryTableM.xml's connectionId attribute.
    Attribute order inside a tag isn't guaranteed by the OOXML spec (real files often put Type before
    Id, e.g.), so every tag is matched whole first and its attributes pulled out independently — never
    assume one attribute precedes another. Any part missing just means that source shows up without a
    sheet, not an error.
    """
    sheet_by_part: dict[str, str] = {}
    try:
        wb_xml = z.read("xl/workbook.xml").decode("utf-8", "ignore")
        rels_xml = z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
        rid_target = {}
        for tag in re.findall(r"<Relationship\b[^>]*/?>", rels_xml):
            rid, target = _attr(tag, "Id"), _attr(tag, "Target")
            if rid and target:
                rid_target[rid] = target
        for tag in re.findall(r"<sheet\b[^>]*/?>", wb_xml):
            name, rid = _attr(tag, "name"), _attr(tag, "r:id")
            target = rid_target.get(rid or "", "")
            if name and target:
                sheet_by_part[os.path.basename(target)] = name
    except (KeyError, OSError):
        return {}

    conn_to_sheet: dict[str, str] = {}
    for item in z.namelist():
        if not (item.startswith("xl/worksheets/_rels/") and item.endswith(".rels")):
            continue
        sheet_name = sheet_by_part.get(item.rsplit("/", 1)[-1][: -len(".rels")])
        if not sheet_name:
            continue
        try:
            raw = z.read(item).decode("utf-8", "ignore")
        except (KeyError, OSError):
            continue
        for tag in re.findall(r"<Relationship\b[^>]*/?>", raw):
            rel_type, target = _attr(tag, "Type") or "", _attr(tag, "Target")
            if "queryTable" not in rel_type or not target:
                continue
            try:
                qt_raw = z.read("xl/queryTables/" + os.path.basename(target)).decode("utf-8", "ignore")
            except (KeyError, OSError):
                continue
            m = re.search(r"<queryTable\b[^>]*/?>", qt_raw)
            cid = _attr(m.group(0), "connectionId") if m else None
            if cid:
                conn_to_sheet[cid] = sheet_name
    return conn_to_sheet


def _external_sources(path: Path) -> list[dict]:
    """
    Best-effort detection of data that enters a workbook through something other than a formula:
    legacy web queries, ODBC/OLEDB/database connections, and Power Query ("Get & Transform") sources.
    Read straight out of the .xlsx zip container — still purely local, still read-only. Each source is
    tagged with the sheet it loads into where that's discoverable (see _connection_sheet_map), so the UI
    can show "this tab is fed by this database" and not just "this file is fed by this database somewhere".

    Power Query's M code is often stored inside a nested, further-compressed "DataMashup" blob, which
    a plain-text regex scan cannot see into — that case is not caught here (see README limits), so those
    sources are reported without a sheet.
    """
    out: list[dict] = []
    try:
        with zipfile.ZipFile(path) as z:
            names = z.namelist()

            if "xl/connections.xml" in names:
                raw = z.read("xl/connections.xml").decode("utf-8", "ignore")
                conn_to_sheet = _connection_sheet_map(z)
                for cid, body in re.findall(r'<connection\b[^>]*\bid="(\d+)"[^>]*>(.*?)</connection>',
                                             raw, re.IGNORECASE | re.DOTALL):
                    sheet = conn_to_sheet.get(cid, "")
                    wm = re.search(r'<webPr\b[^>]*\burl="([^"]+)"', body, re.IGNORECASE)
                    if wm:
                        out.append({"kind": "web_query", "detail": wm.group(1)[:200], "sheet": sheet,
                                    "via": "connections.xml (web query)"})
                    dm = re.search(r'<dbPr\b[^>]*\bconnection="([^"]*)"', body, re.IGNORECASE)
                    if dm:
                        out.append({"kind": "database", "detail": _friendly_conn_detail(dm.group(1)), "sheet": sheet,
                                    "via": "connections.xml (database connection)"})

            # Power Query M formulas: usually in xl/queries/*.xml, sometimes inlined elsewhere under xl/.
            # A plain regex scan for the well-known M "source" functions catches most real-world cases
            # without needing to parse the compressed mashup package. No sheet attribution here — the
            # M code lives outside any one worksheet's XML, unlike a legacy queryTable/webPr/dbPr.
            for item in names:
                if not (item.startswith("xl/queries/") or item.startswith("customXml/")):
                    continue
                try:
                    raw = z.read(item).decode("utf-8", "ignore")
                except (KeyError, OSError):
                    continue
                for m in RE_M_SOURCE.finditer(raw):
                    kind = M_SOURCE_KIND.get(m.group(1).lower(), "web_query")
                    detail = m.group(2).strip()[:200] or "(see Power Query editor for source)"
                    out.append({"kind": kind, "detail": detail, "sheet": "", "via": "Power Query"})
    except (zipfile.BadZipFile, KeyError, OSError):
        pass

    # dedupe identical (kind, detail) pairs found via multiple parts
    seen: set[tuple] = set()
    deduped = []
    for s in out:
        key = (s["kind"], s["detail"])
        if key not in seen:
            seen.add(key)
            deduped.append(s)
    return deduped


def scan_workbook(path: Path, root: Path) -> BookInfo:
    rel = str(path.relative_to(root))
    stat = path.stat()
    import datetime as _dt

    info = BookInfo(
        id=_book_id(rel),
        name=path.name,
        path=str(path),
        rel_path=rel,
        size_kb=max(1, stat.st_size // 1024),
        modified=_dt.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
    )

    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=True, keep_links=True)
    except Exception as exc:  # corrupt / password protected / not really xlsx
        info.error = f"{type(exc).__name__}: {exc}"
        return info

    refs: set[str] = set()
    links: list[dict] = []
    web_links: list[dict] = []
    ext_sources: list[dict] = []

    for ws in wb.worksheets:
        s = SheetInfo(name=ws.title, rows=ws.max_row or 0, cols=ws.max_column or 0)
        header_row: list[str] = []
        header_by_col: dict[int, str] = {}

        for r_i, row in enumerate(ws.iter_rows(max_row=min(ws.max_row or 1, 4000)), start=1):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if r_i == 1 and isinstance(v, str):
                    if len(header_row) < 24:
                        header_row.append(v.strip()[:40])
                    if cell.column and cell.column <= MAX_HEADER_COLS:
                        header_by_col[cell.column] = v.strip()[:40]
                if isinstance(v, str) and v.startswith("="):
                    s.formula_cells += 1
                    if RE_VOLATILE.search(v):
                        s.volatile_cells += 1
                    if RE_LOOKUP.search(v):
                        s.lookup_cells += 1
                    if re.search(r"(?<![A-Z$:!\d])\d{3,}(?![\d:])", v):
                        s.hardcoded_in_formula += 1
                    for m in RE_EXTERNAL.finditer(v):
                        refs.add(m.group(1))
                    if RE_WEBFORMULA.search(v):
                        s.web_formula_cells += 1
                        if len(web_links) < MAX_LINKS_PER_BOOK:
                            urlm = RE_URL.search(v)
                            web_links.append({
                                "dest_sheet": ws.title,
                                "dest_cell": cell.coordinate,
                                "dest_col": get_column_letter(cell.column),
                                "dest_header": header_by_col.get(cell.column, ""),
                                "formula": v[:300],
                                "url": urlm.group(1)[:200] if urlm else "",
                            })
                    if len(links) < MAX_LINKS_PER_BOOK:
                        for m in RE_EXTCELL.finditer(v):
                            src_book = m.group("f1") or m.group("f2")
                            src_sheet = (m.group("s1") or m.group("s2") or "").strip()
                            src_cell = (m.group("c1") or m.group("c2") or "").replace("$", "")
                            links.append({
                                "src_book": src_book,
                                "src_sheet": src_sheet,
                                "src_cell": src_cell,
                                "src_col": _leading_col(src_cell),
                                "dest_sheet": ws.title,
                                "dest_cell": cell.coordinate,
                                "dest_col": get_column_letter(cell.column),
                                "dest_header": header_by_col.get(cell.column, ""),
                                "formula": v[:300],
                            })
                else:
                    s.constant_cells += 1

        s.headers = header_row
        info.sheets.append(s)

    try:
        wb.close()
    except Exception:
        pass

    for t in _external_link_targets(path):
        if t.lower().startswith(("http://", "https://")):
            # a "workbook link" that actually points at a URL, not a local file — e.g. a linked
            # SharePoint/OneDrive workbook or an OLAP cube. Definitely external, definitely not a broken link.
            ext_sources.append({"kind": "web_workbook", "detail": t[:200], "via": "external link (URL target)"})
        elif Path(t).suffix.lower() in EXCEL_EXT:
            refs.add(t)

    ext_sources.extend(_external_sources(path))

    info.external_refs = sorted(refs)
    info.links = links
    info.web_links = web_links
    info.external_sources = ext_sources
    return info


def scan_folder(root: str | Path, max_files: int = 400) -> dict:
    root = Path(root).expanduser().resolve()
    if not root.is_dir():
        raise NotADirectoryError(f"Not a folder: {root}")

    books: list[BookInfo] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if not d.startswith(SKIP_PREFIX)]
        for fn in sorted(filenames):
            if fn.startswith(SKIP_PREFIX):
                continue
            if Path(fn).suffix.lower() not in EXCEL_EXT:
                continue
            books.append(scan_workbook(Path(dirpath) / fn, root))
            if len(books) >= max_files:
                break
        if len(books) >= max_files:
            break

    # resolve external refs to scanned books, by filename
    by_name = {b.name.lower(): b for b in books}
    for b in books:
        for ref in b.external_refs:
            if ref.lower() not in by_name:
                b.unresolved_refs.append(ref)

    return {
        "root": str(root),
        "books": [asdict(b) for b in books],
    }
